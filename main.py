"""Orkestrator: ATC7-universum -> indicaties matchen -> DBC-codes -> output.

Stappen:
1. Universum (in scope) uit Kompas + G-standaard (scripts.drugs).
2. Alle distinctieve indicatieteksten (add-on INKORT + Kompas-bullets) classificeren
   naar een DBC-ziektebeeld of 'geen' (scripts.matching). De LLM-verdicts worden
   gecheckpoint in data/verdicts.json (resumebaar; --reuse-verdicts herbouwt alleen output).
3. Per ATC7 de ziektebeelden bepalen (add-on + Kompas), met off-label-flag en de
   ondersteunende G-standaard- en Kompas-teksten.
4. Wegschrijven: drug_dbc.csv (databron) en dbc_drugs.xlsx voor de apotheker (tab
   "Medicatie" met geel-gemarkeerde te-checken koppelingen + tab per ziektebeeld). De
   apotheker corrigeert in de kolom "correctie" en draait de macro
   macro/HergroepeerZiektebeelden.bas om de ziektebeeld-tabbladen te hergroeperen.

Draaien: ./venv/bin/python main.py --model qwen2.5:3b-instruct
Opties: --no-llm, --no-embeddings, --reuse-verdicts, --model, --limit N
"""
from __future__ import annotations

import argparse
import collections
import csv
import json
from pathlib import Path

from openpyxl import Workbook

from scripts import drugs
from scripts.dbc import load_ziektebeelden
from scripts.diff_oude_lijst import bereken_diff
from scripts.gstandaard import DEFAULT_DB_PATH, GstandaardDB
from scripts.matching import GEEN, OLLAMA_MODEL, Matcher, Verdict

_ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = _ROOT / "output"
VERDICTS_CACHE = _ROOT / "data" / "verdicts.json"


def _categorie(g) -> str:
    """Vergoedingscategorie van het middel: Add-on (G-standaard BST131) of EVS."""
    return "Add-on" if g.addon else "EVS"


def _atc_groep_ziektebeelden(atc7: str) -> list[str]:
    """Directe ATC-groep-regel voor diabetes/obesitas (geen tekst-matching nodig).

    A10 (incl. GLP-1's A10BJ) = diabetes, A08 = antiobesitas, exact volgens de ATC-indeling.
    Oncologie (L) valt hier bewust buiten: die groep zegt niets over de tumor.
    """
    if atc7.startswith("A08"):
        return ["obesitas"]
    if atc7.startswith("A10"):
        return ["diabetes"]
    return []


# -- classificatie ---------------------------------------------------------

def _verzamel_teksten(universe) -> list[str]:
    teksten: list[str] = []
    for g in universe.values():
        teksten += [i["inkort"] for i in g.addon]
        teksten += g.kompas_indicaties
    return list(dict.fromkeys(teksten))


def _classificeer(universe, ziektebeelden, gebruik_embeddings, gebruik_llm, model, reuse,
                  verifieer_lexicon=False) -> dict[str, Verdict]:
    teksten = _verzamel_teksten(universe)
    if reuse:
        cache = json.loads(VERDICTS_CACHE.read_text(encoding="utf-8")) if VERDICTS_CACHE.is_file() else {}
        ontbreekt = sum(1 for t in teksten if t not in cache)
        print(f"  verdicts uit cache ({VERDICTS_CACHE.name}); {ontbreekt} niet in cache -> geen")
        return {t: Verdict(**cache[t]) if t in cache else Verdict(t, GEEN, 0.0, "geen-default")
                for t in teksten}
    matcher = Matcher(ziektebeelden, gebruik_embeddings=gebruik_embeddings,
                      gebruik_llm=gebruik_llm, ollama_model=model, verifieer_lexicon=verifieer_lexicon)
    print(f"  matcher: embeddings={'aan' if matcher._embedder else 'uit'} "
          f"llm={'aan' if matcher.llm else 'uit'} | {len(teksten)} distinct teksten")
    return matcher.classify_many(teksten, checkpoint_path=VERDICTS_CACHE)


# -- koppelingen -----------------------------------------------------------

def _te_checken(info) -> bool:
    """Geel = de apotheker moet kijken: off-label, een LLM-beslissing, of (met
    --verify-lexicon) een LLM die van het lexicon afweek. Lexicon/ATC-groep/geen = niet geel."""
    if info["methode"] == "geen":
        return False
    if info["off_label"] or info["flag"]:
        return True
    return info["methode"] not in ("lexicon", "atc-groep", "handmatig")


def _geen_rij(g) -> dict:
    """Koppelinfo voor een middel zonder ziektebeeld, zodat het toch in de Medicatie-tab
    staat en de apotheker het via 'correctie' alsnog aan een ziektebeeld kan toewijzen."""
    return {"bron": "", "off_label": False, "confidence": 0.0, "methode": "geen", "flag": False,
            "lexicon": "", "gstandaard": " | ".join(i["inkort"] for i in g.addon),
            "kompas": " | ".join(g.kompas_indicaties)}


def _ziektebeelden_per_atc(g, verdicts) -> dict[str, dict]:
    """Per ATC7: ziektebeeld -> koppelinfo met bron, off-label-flag, herkomstteksten.

    Verzamelt per ziektebeeld ALLE ondersteunende add-on- (G-standaard) en Kompas-teksten;
    de hoogste-confidence indicatie bepaalt methode/flag/lexicon. off_label = de koppeling
    rust uitsluitend op off-label add-on (geen geregistreerde add-on, geen Kompas).
    """
    acc: dict[str, dict] = {}

    def vak(zb):
        return acc.setdefault(zb, {"gstandaard": [], "kompas": [], "aard": set(),
                                   "best": -1.0, "methode": "", "flag": False, "lexicon": ""})

    def update_best(info, v):
        if v.score > info["best"]:
            info["best"], info["methode"], info["flag"], info["lexicon"] = v.score, v.methode, v.flag, v.lexicon

    for ind in g.addon:
        v = verdicts[ind["inkort"]]
        if v.ziektebeeld == GEEN:
            continue
        info = vak(v.ziektebeeld)
        info["gstandaard"].append((ind["inkort"], ind["indicatie_aard"]))
        info["aard"].add("offlabel" if "off-label" in ind["indicatie_aard"] else "geregistreerd")
        update_best(info, v)
    for bullet in g.kompas_indicaties:
        v = verdicts[bullet]
        if v.ziektebeeld == GEEN:
            continue
        info = vak(v.ziektebeeld)
        info["kompas"].append(bullet)
        update_best(info, v)

    uit = {}
    for zb, info in acc.items():
        uit[zb] = {
            "bron": "add-on" if info["gstandaard"] else "kompas",
            "off_label": info["aard"] == {"offlabel"} and not info["kompas"],
            "confidence": info["best"], "methode": info["methode"], "flag": info["flag"],
            "lexicon": info["lexicon"],
            "gstandaard": " | ".join(t for t, _ in info["gstandaard"]),
            "kompas": " | ".join(info["kompas"]),
        }
    return uit


def _atc_groep_koppeling(zbs) -> dict[str, dict]:
    return {zb: {"bron": "atc-groep", "off_label": False, "confidence": 1.0, "methode": "atc-groep",
                 "flag": False, "lexicon": "", "gstandaard": "", "kompas": ""} for zb in zbs}


# -- output ----------------------------------------------------------------

def _schrijf_drug_dbc(universe, koppelingen, ziektebeelden):
    pad = OUTPUT_DIR / "drug_dbc.csv"
    with open(pad, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["atc7", "stofnaam", "ziektebeeld", "bron", "off_label", "te_checken",
                    "confidence", "methode", "lexicon_suggestie", "gstandaard_indicatie",
                    "kompas_indicatie", "specialisme_code", "diagnose_code", "omschrijving"])
        for atc7, g in sorted(universe.items()):
            for slug, info in sorted(koppelingen[atc7].items()):
                codes = ziektebeelden[slug].dbc_codes if slug in ziektebeelden else [None]
                for code in codes:
                    w.writerow([atc7, g.stofnaam, slug, info["bron"],
                                "ja" if info["off_label"] else "", "ja" if _te_checken(info) else "",
                                f"{info['confidence']:.2f}", info["methode"], info["lexicon"],
                                info["gstandaard"], info["kompas"],
                                code.specialisme_code if code else "",
                                code.diagnose_code if code else "",
                                code.omschrijving if code else ""])
    return pad


def _schrijf_deliverable_xlsx(universe, koppelingen, ziektebeelden, slugs, diff):
    """Excel voor de apotheker: tab "Medicatie" (geel = te checken) + tab per ziektebeeld.

    Per ziektebeeld-tab een diff-sectie t.o.v. de oude DMA-lijst (nieuw/verwijderd/
    gebleven) en een samenvattingstab "Diff-overzicht" met de aantallen. Een correctie in
    de dropdown van "Medicatie" gaat via apply_review naar de overrides en hergroepeert
    bij een herdraai automatisch ook de ziektebeeld-tabs.
    """
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    GEEL = PatternFill("solid", fgColor="FFF2A8")
    TEAL = PatternFill("solid", fgColor="0E7C86")
    KOP = Font(bold=True, color="FFFFFF")
    VET = Font(bold=True)

    def kleur_kop(ws, rij=1):
        for c in ws[rij]:
            c.font, c.fill = KOP, TEAL

    def geel(ws):
        for c in ws[ws.max_row]:
            c.fill = GEEL

    wb = Workbook()

    # --- Tab 1: Medicatie ---
    ws = wb.active
    ws.title = "Medicatie"
    kop = ["ziektebeeld", "atc7", "stofnaam", "categorie", "off_label", "zekerheid",
           "methode", "lexicon_suggestie", "gstandaard_indicatie", "kompas_indicatie", "correctie"]
    ws.append(kop)
    rijen = []
    for atc7, kopp in koppelingen.items():
        for zb, info in kopp.items():
            rijen.append((atc7, zb, info))
    # te-checken (geel) eerst, dan de rest, en de 'geen'-rijen helemaal onderaan.
    rijen.sort(key=lambda r: (r[2]["methode"] == "geen", not _te_checken(r[2]), r[1], r[0]))
    for atc7, zb, info in rijen:
        g = universe[atc7]
        ws.append([zb, atc7, g.stofnaam, _categorie(g), "ja" if info["off_label"] else "",
                   round(info["confidence"], 2), info["methode"], info["lexicon"],
                   info["gstandaard"], info["kompas"], ""])
        if _te_checken(info):
            geel(ws)
    kleur_kop(ws)
    dv = DataValidation(type="list", formula1=f'"{",".join(list(slugs) + ["geen"])}"', allow_blank=True,
                        showInputMessage=True, promptTitle="Correctie",
                        prompt="Kies een ziektebeeld of 'geen'. Leeg laten = voorstel akkoord.")
    ws.add_data_validation(dv)
    dv.add(f"K2:K{ws.max_row}")
    for col, br in {"A": 20, "B": 9, "C": 26, "D": 9, "E": 9, "F": 9, "G": 11, "H": 16,
                    "I": 50, "J": 50, "K": 20}.items():
        ws.column_dimensions[col].width = br
    ws.freeze_panes = "A2"

    # --- Tab per ziektebeeld ---
    for slug, z in ziektebeelden.items():
        ws = wb.create_sheet(title=slug[:31])
        ws.append(["DBC-codes"]); ws[ws.max_row][0].font = VET
        ws.append(["specialisme_code", "diagnose_code", "omschrijving"]); kleur_kop(ws, ws.max_row)
        for c in z.dbc_codes:
            ws.append([c.specialisme_code, c.diagnose_code, c.omschrijving])
        ws.append([]); ws.append(["Geneesmiddelen"]); ws[ws.max_row][0].font = VET
        ws.append(["atc7", "stofnaam", "categorie", "off_label", "zekerheid", "methode"])
        kleur_kop(ws, ws.max_row)
        meds = sorted((a for a, k in koppelingen.items() if slug in k),
                      key=lambda a: (not _te_checken(koppelingen[a][slug]), a))
        for atc7 in meds:
            info = koppelingen[atc7][slug]
            ws.append([atc7, universe[atc7].stofnaam, _categorie(universe[atc7]),
                       "ja" if info["off_label"] else "", round(info["confidence"], 2), info["methode"]])
            if _te_checken(info):
                geel(ws)

        # Diff t.o.v. de oude DMA-lijst
        d = diff.get(slug, {"nieuw": [], "verwijderd": [], "gebleven": [], "buiten_scope": []})
        ws.append([]); ws.append(["Diff t.o.v. oude lijst"]); ws[ws.max_row][0].font = VET
        verw_evs = sum(1 for *_, h in d["verwijderd"] if h == "EVS")
        ws.append([f"nieuw: {len(d['nieuw'])}  |  verwijderd: {len(d['verwijderd'])} "
                   f"(EVS: {verw_evs}, add-on: {len(d['verwijderd']) - verw_evs})  |  "
                   f"gebleven: {len(d['gebleven'])}"])
        ws.append(["status", "atc7", "stofnaam", "oude_herkomst"]); kleur_kop(ws, ws.max_row)
        for status, items in (("nieuw", d["nieuw"]), ("verwijderd", d["verwijderd"]),
                              ("gebleven", d["gebleven"])):
            for atc7, naam, herkomst in items:
                ws.append([status, atc7, naam, herkomst])
        if d["buiten_scope"]:
            ws.append([]); ws.append(["Oude EVS-codes buiten onze scope (niet vergeleken)"])
            ws[ws.max_row][0].font = VET
            for atc7, naam in d["buiten_scope"]:
                ws.append(["buiten-scope", atc7, naam, "EVS"])
        for col, br in {"A": 13, "B": 10, "C": 26, "D": 9, "E": 10, "F": 12}.items():
            ws.column_dimensions[col].width = br

    # --- Tab Diff-overzicht ---
    ws = wb.create_sheet(title="Diff-overzicht")
    ws.append(["ziektebeeld", "nieuw", "verwijderd", "verw_EVS", "verw_addon",
               "gebleven", "buiten_scope"])
    kleur_kop(ws)
    for slug in ziektebeelden:
        d = diff.get(slug, {})
        verw = d.get("verwijderd", [])
        verw_evs = sum(1 for *_, h in verw if h == "EVS")
        ws.append([slug, len(d.get("nieuw", [])), len(verw), verw_evs, len(verw) - verw_evs,
                   len(d.get("gebleven", [])), len(d.get("buiten_scope", []))])
    for col, br in {"A": 22, "B": 9, "C": 11, "D": 10, "E": 11, "F": 10, "G": 13}.items():
        ws.column_dimensions[col].width = br
    ws.freeze_panes = "A2"

    pad = OUTPUT_DIR / "dbc_drugs.xlsx"
    wb.save(pad)
    return pad


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--no-embeddings", action="store_true")
    ap.add_argument("--no-llm", action="store_true")
    ap.add_argument("--model", default=OLLAMA_MODEL, help="Ollama-model voor de LLM-laag")
    ap.add_argument("--reuse-verdicts", action="store_true",
                    help="laad de classificatie uit data/verdicts.json (geen LLM); alleen output herbouwen")
    ap.add_argument("--verify-lexicon", action="store_true",
                    help="laat de LLM ook lexicon-treffers herbeoordelen (alleen met een sterk model, bv. 14B)")
    ap.add_argument("--limit", type=int, default=0, help="beperk universum (test)")
    args = ap.parse_args()

    if not DEFAULT_DB_PATH.is_file():
        raise SystemExit("Bouw eerst de SQLite-cache: ./venv/bin/python -m scripts.build_gstandaard_db")

    print("Universum opbouwen...")
    universe = drugs.build_universe(db=GstandaardDB())
    if args.limit:
        universe = dict(list(universe.items())[: args.limit])
    print(f"  {len(universe)} ATC7 in scope")

    ziektebeelden = load_ziektebeelden()
    via_engine = {a: g for a, g in universe.items() if not _atc_groep_ziektebeelden(a)}
    print(f"  {len(universe) - len(via_engine)} ATC7 direct via ATC-groep (diabetes/obesitas), "
          f"{len(via_engine)} via de matching-engine (oncologie)")

    print("Indicaties classificeren...")
    verdicts = _classificeer(via_engine, ziektebeelden,
                             gebruik_embeddings=not args.no_embeddings, gebruik_llm=not args.no_llm,
                             model=args.model, reuse=args.reuse_verdicts,
                             verifieer_lexicon=args.verify_lexicon)

    koppelingen = {
        atc7: (_atc_groep_koppeling(groep) if (groep := _atc_groep_ziektebeelden(atc7))
               else _ziektebeelden_per_atc(g, verdicts))
        for atc7, g in universe.items()
    }
    # Middelen zonder ziektebeeld krijgen een 'geen'-rij (rescue via de correctie-dropdown).
    for atc7, g in universe.items():
        if not koppelingen[atc7]:
            koppelingen[atc7]["geen"] = _geen_rij(g)

    nieuw_per_slug: dict[str, set] = collections.defaultdict(set)
    for atc7, kopp in koppelingen.items():
        for slug in kopp:
            if slug != GEEN:
                nieuw_per_slug[slug].add(atc7)
    diff = bereken_diff(nieuw_per_slug, universe)

    OUTPUT_DIR.mkdir(exist_ok=True)
    p1 = _schrijf_drug_dbc(universe, koppelingen, ziektebeelden)
    p2 = _schrijf_deliverable_xlsx(universe, koppelingen, ziektebeelden, list(ziektebeelden), diff)

    per_zb = collections.Counter(slug for k in koppelingen.values() for slug in k if slug != GEEN)
    te_checken = sum(1 for k in koppelingen.values() for info in k.values() if _te_checken(info))
    gekoppeld = sum(1 for k in koppelingen.values() if any(zb != GEEN for zb in k))
    print(f"Klaar. {gekoppeld}/{len(universe)} ATC7 gekoppeld aan >=1 ziektebeeld; "
          f"{len(universe) - gekoppeld} zonder (geen); {te_checken} koppelingen te checken (geel).")
    print(f"  geneesmiddelen per ziektebeeld: {dict(per_zb.most_common())}")
    for p in (p1, p2):
        print(f"  -> {p}")


if __name__ == "__main__":
    main()
