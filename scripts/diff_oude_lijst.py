"""Diff van de nieuwe koppeling t.o.v. de oude DMA-lijst (per ziektebeeld).

De oude lijst (`data/ATC inclusion DMA.xlsx`) heeft een tabblad per ziektebeeld met
twee mogelijke secties: `add-on exclusion` (ATC-codes die NIET meetellen, gemengde
granulariteit: niveau-1 letter, niveau-4, of 7-char) en `EVS inclusion` (expliciet
wel-meegerekende 7-char codes). De nieuwe lijst is inclusie-gebaseerd (per ATC7 de
gekoppelde ziektebeelden), dus de oude exclusie wordt binnen onze scope geinverteerd:
een in-scope add-on-middel is oud-inbegrepen tenzij een exclusie-prefix het dekt.

Sectiekoppen worden op substring herkend (`excl`/`incl`) zodat typefouten in het
bronbestand (`exlcusion`, `addon - exclusion`) geen rol spelen.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

_ROOT = Path(__file__).resolve().parent.parent
OUDE_LIJST_PAD = _ROOT / "data" / "ATC inclusion DMA.xlsx"

# Tabblad-afkorting -> ziektebeeld-slug. Vastgesteld via de medicatie-inhoud:
# CRC sluit cisplatine/paclitaxel uit en houdt oxaliplatine/irinotecan/capecitabine in
# (colorectaal); CC houdt die gynae-middelen in (cervix); VC = vulva.
SHEET_NAAR_SLUG = {
    "BC": "borstkanker",
    "CRC": "darmkanker",
    "LC": "longkanker",
    "PC": "prostaatkanker",
    "OC": "ovariumcarcinoom",
    "VC": "vulvacarcinoom",
    "EC": "endometriumcarcinoom",
    "CC": "cervixcarcinoom",
    "GC": "maag-slokdarmkanker",
    "HN": "hoofd-halskanker",
    "Panc": "alvleesklierkanker",
    "DM": "diabetes",
    "MC": "melanoom",
    "OBE": "obesitas",
}

# Geldige ATC-token: hoofdletter gevolgd door alleen cijfers/hoofdletters. Vangt de
# niveau-1 letters (A) tot en met 7-char codes (L01XX35) en weert lege cellen, de
# boolean-glitch (False) en beschrijvingen (kleine letters / spaties).
_ATC = re.compile(r"^[A-Z][0-9A-Z]*$")


def _is_atc(v) -> bool:
    return not isinstance(v, bool) and isinstance(v, str) and bool(_ATC.match(v.strip()))


def _code_kolommen(rows) -> list[tuple[int, str]]:
    """Bepaal per `code`-kolom het sectietype ('excl'/'incl') uit de eerste twee rijen."""
    koppen = rows[0] if rows else ()
    subkoppen = rows[1] if len(rows) > 1 else ()

    secties = []  # (kolomindex, type) van elke sectiekop, op volgorde
    for idx, cel in enumerate(koppen):
        if not isinstance(cel, str):
            continue
        laag = cel.lower()
        # Inclusie eerst; exclusie tolereert de bron-typefout 'exlcusion' (c/l verwisseld).
        if "incl" in laag:
            secties.append((idx, "incl"))
        elif "exc" in laag or "exl" in laag:
            secties.append((idx, "excl"))

    uit = []
    for idx, cel in enumerate(subkoppen):
        if isinstance(cel, str) and cel.strip().lower() == "code":
            horend = [(s_idx, s_type) for s_idx, s_type in secties if s_idx <= idx]
            if horend:
                uit.append((idx, horend[-1][1]))
    return uit


def parse_oude_lijst(pad: Path | str = OUDE_LIJST_PAD) -> dict[str, dict]:
    """Per ziektebeeld-slug: {'exclusie': set[code], 'evs': {code: omschrijving},
    'heeft_exclusie': bool}. Bladen die op dezelfde slug mappen worden samengevoegd.
    """
    wb = openpyxl.load_workbook(pad, data_only=True, read_only=True)
    per_slug: dict[str, dict] = {}
    for blad in wb.sheetnames:
        slug = SHEET_NAAR_SLUG.get(blad)
        if slug is None:
            continue
        rows = list(wb[blad].iter_rows(values_only=True))
        acc = per_slug.setdefault(slug, {"exclusie": set(), "evs": {}, "heeft_exclusie": False})
        for kol, soort in _code_kolommen(rows):
            if soort == "excl":
                acc["heeft_exclusie"] = True
            for r in rows[2:]:
                code = r[kol] if kol < len(r) else None
                if not _is_atc(code):
                    continue
                code = code.strip()
                if soort == "excl":
                    acc["exclusie"].add(code)
                else:
                    desc = r[kol + 1] if kol + 1 < len(r) and isinstance(r[kol + 1], str) else ""
                    acc["evs"].setdefault(code, desc.strip() if desc else "")
    wb.close()
    return per_slug


def oude_inclusie_per_slug(parsed: dict[str, dict], universe) -> dict[str, dict]:
    """Inverteer de exclusie binnen scope -> per slug de oud-inbegrepen ATC7's met herkomst.

    Een in-scope add-on-middel is oud-inbegrepen ('add-on') als geen exclusie-prefix het
    dekt (alleen wanneer het blad een exclusie-sectie had). EVS-codes worden toegevoegd
    ('EVS', wint van de inversie). EVS-codes buiten onze scope komen apart terug als
    'buiten_scope'. 'inclusie' = {atc7: herkomst}.
    """
    addon_scope = [atc7 for atc7, g in universe.items() if g.addon]
    uit: dict[str, dict] = {}
    for slug, acc in parsed.items():
        excl = acc["exclusie"]
        herkomst: dict[str, str] = {}
        if acc["heeft_exclusie"]:
            for a in addon_scope:
                if not any(a.startswith(e) for e in excl):
                    herkomst[a] = "add-on"
        for c in acc["evs"]:
            if c in universe:
                herkomst[c] = "EVS"
        buiten = {c: d for c, d in acc["evs"].items() if c not in universe}
        uit[slug] = {"inclusie": herkomst, "buiten_scope": buiten}
    return uit


def _rij(codes, universe, herkomst=None) -> list[tuple[str, str, str]]:
    """(atc7, stofnaam, herkomst) gesorteerd op atc7. herkomst-dict mapt atc7->bron;
    ontbreekt die (nieuwe codes), dan '-'."""
    herkomst = herkomst or {}
    uit = []
    for c in sorted(codes):
        naam = universe[c].stofnaam if c in universe else ""
        uit.append((c, naam, herkomst.get(c, "-")))
    return uit


def diff_per_slug(oude: dict[str, dict], nieuw_per_slug: dict[str, set], universe) -> dict[str, dict]:
    """Per slug: nieuw / verwijderd / gebleven (ATC7 + stofnaam + herkomst) + buiten-scope."""
    slugs = set(oude) | set(nieuw_per_slug)
    uit: dict[str, dict] = {}
    for slug in slugs:
        oud = oude.get(slug, {}).get("inclusie", {})   # {atc7: herkomst}
        oud_set = set(oud)
        new_set = set(nieuw_per_slug.get(slug, set()))
        buiten = oude.get(slug, {}).get("buiten_scope", {})
        uit[slug] = {
            "nieuw": _rij(new_set - oud_set, universe),
            "verwijderd": _rij(oud_set - new_set, universe, oud),
            "gebleven": _rij(new_set & oud_set, universe, oud),
            "buiten_scope": [(c, buiten[c]) for c in sorted(buiten)],
        }
    return uit


def bereken_diff(nieuw_per_slug: dict[str, set], universe,
                 pad: Path | str = OUDE_LIJST_PAD) -> dict[str, dict]:
    """Volledige keten: oude lijst inlezen, inverteren, diffen tegen de nieuwe koppeling."""
    parsed = parse_oude_lijst(pad)
    oude = oude_inclusie_per_slug(parsed, universe)
    return diff_per_slug(oude, nieuw_per_slug, universe)
