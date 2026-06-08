"""Pas de HITL-correcties uit output/dbc_drugs.xlsx (tab Controle) toe op data/overrides.json.

Werkwijze:
1. Draai `main.py` -> schrijft output/dbc_drugs.xlsx (tab "Controle", kolom `correctie`).
2. Vul in de kolom `correctie` het juiste ziektebeeld in waar het voorstel fout is
   (leeg laten = voorstel accepteren).
3. Draai dit script -> de ingevulde correcties gaan naar data/overrides.json.
4. Draai `main.py --reuse-verdicts` opnieuw -> de overrides worden toegepast (geen LLM nodig).

    ./venv/bin/python -m scripts.apply_review
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

_ROOT = Path(__file__).resolve().parent.parent
DELIVERABLE = _ROOT / "output" / "dbc_drugs.xlsx"
OVERRIDES = _ROOT / "data" / "overrides.json"


def apply() -> None:
    if not DELIVERABLE.is_file():
        raise SystemExit(f"{DELIVERABLE} ontbreekt; draai eerst main.py")
    wb = load_workbook(DELIVERABLE, read_only=True)
    ws = wb["Medicatie"]
    kop = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    i_tekst, i_corr = kop.index("onderbouwing"), kop.index("correctie")

    overrides = json.loads(OVERRIDES.read_text(encoding="utf-8")) if OVERRIDES.is_file() else {}
    toegevoegd = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        tekst, corr = row[i_tekst], row[i_corr]
        if tekst and corr and str(corr).strip():
            overrides[str(tekst).lower()] = str(corr).strip()
            toegevoegd += 1
    OVERRIDES.parent.mkdir(parents=True, exist_ok=True)
    OVERRIDES.write_text(json.dumps(overrides, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"{toegevoegd} correcties verwerkt -> {OVERRIDES} ({len(overrides)} overrides totaal)")


if __name__ == "__main__":
    apply()
